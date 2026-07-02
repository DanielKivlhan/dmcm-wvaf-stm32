"""
generate_excel_data.py
======================
Membuat file Excel (dmcm_lm35_data.xlsx) dari lm35_clean.dat
untuk digunakan sebagai data GNU Plot dan dibuka di spreadsheet.

Kolom:
  A: index     - Nomor sampel
  B: time_s    - Waktu (detik)
  C: T_raw     - Suhu mentah LM35 (derajat C)
  D: T_filtered- Suhu setelah WVAF filter (derajat C)
  E: N_win     - Ukuran window adaptif
  F: DeltaT    - |DeltaT| Rate-of-change (derajat C)
  G: State     - State mesin (0=Normal, 1=Warning, 2=Danger)
  H: Relay     - Output relay (0=Off, 1=On)
  I: LED_Red   - LED merah (0=Off, 1=On)
  J: LED_Green - LED hijau (0=Off, 1=On)
"""

import os
import csv

# -- path relatif ke direktori script ini
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
INPUT_DAT  = os.path.join(BASE_DIR, "lm35_clean.dat")
OUTPUT_XLS = os.path.join(BASE_DIR, "dmcm_lm35_data.xlsx")
OUTPUT_CSV = os.path.join(BASE_DIR, "dmcm_lm35_data.csv")

# -- baca data
rows = []
with open(INPUT_DAT, "r") as f:
    reader = csv.reader(f)
    for i, row in enumerate(reader):
        if not row or row[0].strip() == "":
            continue
        rows.append(row)

header = rows[0]
data   = rows[1:]

# -- header tampilan
HEADER_DISPLAY = [
    "index", "time_s (s)", "T_raw (deg C)", "T_filtered (deg C)",
    "N_win", "|DeltaT| (deg C)", "State (0=N,1=W,2=D)", "Relay", "LED_Red", "LED_Green"
]

# -- tulis CSV biasa (bisa dibuka Excel/LibreOffice)
with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(HEADER_DISPLAY)
    for row in data:
        if len(row) >= 10:
            writer.writerow([
                int(row[0]),
                float(row[1]),
                float(row[2]),
                float(row[3]),
                int(row[4]),
                float(row[5]),
                int(row[6]),
                int(row[7]),
                int(row[8]),
                int(row[9]),
            ])

print(f"[OK] CSV ditulis  : {OUTPUT_CSV}")

# -- coba buat Excel jika openpyxl tersedia
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DMCM-LM35 Data"

    # warna header (biru IEEE)
    HDR_FILL  = PatternFill("solid", fgColor="1A5ECC")
    HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    CELL_FONT = Font(name="Arial", size=9)
    THIN      = Side(style="thin", color="CCCCCC")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER    = Alignment(horizontal="center", vertical="center")

    # kelompok warna baris berdasarkan state
    STATE_FILL = {
        0: PatternFill("solid", fgColor="E8F5E9"),   # Normal  - hijau muda
        1: PatternFill("solid", fgColor="FFF8E1"),   # Warning - kuning muda
        2: PatternFill("solid", fgColor="FFEBEE"),   # Danger  - merah muda
    }

    # tulis header
    for col_idx, col_name in enumerate(HEADER_DISPLAY, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        cell.border    = BORDER

    # tulis data baris per baris
    for row_idx, row in enumerate(data, start=2):
        if len(row) < 10:
            continue
        state = int(row[6])
        row_fill = STATE_FILL.get(state, PatternFill())

        values = [
            int(row[0]),
            float(row[1]),
            float(row[2]),
            float(row[3]),
            int(row[4]),
            float(row[5]),
            int(row[6]),
            int(row[7]),
            int(row[8]),
            int(row[9]),
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = CELL_FONT
            cell.fill      = row_fill
            cell.border    = BORDER
            cell.alignment = CENTER

    # lebar kolom
    col_widths = [8, 12, 15, 17, 8, 16, 22, 8, 10, 11]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # freeze baris header
    ws.freeze_panes = "A2"

    # tambah sheet keterangan
    ws_info = wb.create_sheet("Keterangan")
    info = [
        ["DMCM-WVAF LM35 Monitoring - Keterangan Data", "", ""],
        ["", "", ""],
        ["Kolom",  "Nama Variabel",        "Keterangan"],
        ["A",      "index",                "Nomor sampel (0 - 512)"],
        ["B",      "time_s",               "Waktu sampling (detik, interval 0.2s)"],
        ["C",      "T_raw",                "Suhu mentah dari ADC LM35 (deg C)"],
        ["D",      "T_filtered",           "Suhu setelah WVAF low-pass filter (deg C)"],
        ["E",      "N_win",                "Ukuran window adaptif WVAF (4 - 8)"],
        ["F",      "|DeltaT|",             "Rate-of-change suhu antar sampel (deg C)"],
        ["G",      "State",               "0=Normal, 1=Warning (>=30C), 2=Danger (>=40C)"],
        ["H",      "Relay",               "Output relay pendingin (0=Off, 1=On)"],
        ["I",      "LED_Red",             "LED merah indikator bahaya (0=Off, 1=On)"],
        ["J",      "LED_Green",           "LED hijau indikator normal (0=Off, 1=On)"],
        ["", "", ""],
        ["Parameter",   "Nilai",           "Keterangan"],
        ["T_warn",      "30 deg C",        "Batas suhu Warning"],
        ["T_crit",      "40 deg C",        "Batas suhu Danger / relay aktif"],
        ["Sampling",    "fs = 5 Hz",       "Interval 0.2 detik per sampel"],
        ["Durasi",      "102.4 detik",     "Total 512 sampel"],
        ["", "", ""],
        ["File GNU Plot",  "lm35_clean.dat",       "Data file untuk GNU Plot (separator koma)"],
        ["Script PNG",     "plot_lm35_ieee.gp",    "Script GNU Plot IEEE style (PNG + PDF)"],
        ["Script Baru",    "plot_lm35_excel.gp",   "Script GNU Plot menggunakan dmcm_lm35_data.csv"],
    ]

    for r, row_data in enumerate(info, start=1):
        for c, val in enumerate(row_data, start=1):
            cell = ws_info.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(name="Arial", bold=True, size=12, color="1A5ECC")
            elif r in (3, 15):
                cell.font = Font(name="Arial", bold=True, size=10)
                cell.fill = PatternFill("solid", fgColor="E3EAF8")
            else:
                cell.font = Font(name="Arial", size=9)

    ws_info.column_dimensions["A"].width = 18
    ws_info.column_dimensions["B"].width = 28
    ws_info.column_dimensions["C"].width = 48

    # Warna state legend
    ws_legend = wb.create_sheet("Warna State")
    legend_data = [
        ["State", "Nilai", "Warna Baris", "Kondisi"],
        ["Normal",  "0", "Hijau muda",  "T < 30 deg C"],
        ["Warning", "1", "Kuning muda", "30 <= T < 40 deg C"],
        ["Danger",  "2", "Merah muda",  "T >= 40 deg C"],
    ]
    legend_fills = [
        PatternFill("solid", fgColor="1A5ECC"),
        PatternFill("solid", fgColor="E8F5E9"),
        PatternFill("solid", fgColor="FFF8E1"),
        PatternFill("solid", fgColor="FFEBEE"),
    ]
    legend_fonts = [
        Font(name="Arial", bold=True, color="FFFFFF", size=10),
        Font(name="Arial", size=10),
        Font(name="Arial", size=10),
        Font(name="Arial", size=10),
    ]
    for r, (row_data, fill, font) in enumerate(zip(legend_data, legend_fills, legend_fonts), start=1):
        for c, val in enumerate(row_data, start=1):
            cell = ws_legend.cell(row=r, column=c, value=val)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
    for col_ltr, w in zip(["A","B","C","D"], [12, 8, 16, 22]):
        ws_legend.column_dimensions[col_ltr].width = w

    wb.save(OUTPUT_XLS)
    print(f"[OK] Excel ditulis: {OUTPUT_XLS}")

except ImportError:
    print("[!] openpyxl tidak tersedia -- hanya CSV yang ditulis.")
    print("    Install dengan: pip install openpyxl")
    print(f"    CSV sudah tersedia di: {OUTPUT_CSV}")

print("\nSelesai!")
print(f"  CSV  -> {OUTPUT_CSV}")
print(f"  XLSX -> {OUTPUT_XLS}")
